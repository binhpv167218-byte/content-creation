#!/usr/bin/env python3
import json, io, warnings
warnings.filterwarnings('ignore')

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import openpyxl
from openpyxl.cell.cell import MergedCell

CREDS_FILE     = '/Users/macos/.config/gdrive-rw-token.json'
KEYS_FILE      = '/Users/macos/.npm-global/lib/node_modules/gcp-oauth.keys.json'
SPREADSHEET_ID = '1FbUqSERC8OCxjKk5zC-962BHmIV7fepp'

with open(CREDS_FILE) as f: token = json.load(f)
with open(KEYS_FILE) as f: keys = json.load(f)
key_type = list(keys.keys())[0]
creds = Credentials(
    token=token['access_token'], refresh_token=token['refresh_token'],
    token_uri=token.get('token_uri', keys[key_type]['token_uri']),
    client_id=token.get('client_id', keys[key_type]['client_id']),
    client_secret=token.get('client_secret', keys[key_type]['client_secret']),
    scopes=token.get('scopes', token.get('scope', '').split()),
)
if creds.expired: creds.refresh(Request())
drive = build('drive', 'v3', credentials=creds)

def write_cell(ws, coord, value):
    """Ghi vào ô, xử lý cả merged cell."""
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        # Tìm master cell của merged range
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                master = ws.cell(rng.min_row, rng.min_col)
                master.value = value
                return master.coordinate
        return None
    else:
        cell.value = value
        return coord

# Download
print("Downloading...")
buf = io.BytesIO()
downloader = MediaIoBaseDownload(buf, drive.files().get_media(fileId=SPREADSHEET_ID))
done = False
while not done: _, done = downloader.next_chunk()

buf.seek(0)
wb = openpyxl.load_workbook(buf)
print("Sheets:", wb.sheetnames)

# Cells cần sửa (đã xác nhận từ lần chạy trước)
# G11 = Giá niêm yết input (= 3,878,116,343 → TỔNG Không Vay = 3,500,000,000)
EDITS = [('G5', 'S5B'), ('G6', 15), ('G7', '08'), ('G8', 'S5B-15-08'), ('G9', '1BR+'), ('G10', 50), ('G11', 3_878_116_343)]

for ws in wb.worksheets:
    for coord, val in EDITS:
        result = write_cell(ws, coord, val)
        print(f"  [{ws.title}] {coord} → {val} (wrote to {result})")

# Upload lại
print("\nUploading...")
out = io.BytesIO()
wb.save(out)
out.seek(0)
mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
drive.files().update(
    fileId=SPREADSHEET_ID,
    media_body=MediaIoBaseUpload(out, mimetype=mime, resumable=True)
).execute()

print("Xong! Mở Google Drive để kiểm tra.")
