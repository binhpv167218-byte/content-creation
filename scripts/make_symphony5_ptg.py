#!/usr/bin/env python3
"""Tạo Phiếu Tính Giá Symphony 5 — template điền giá là ra kết quả"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Màu sắc ──────────────────────────────────────────────
C_NAVY   = "1A3A5C"   # header chính
C_ORANGE = "E87722"   # accent
C_GOLD   = "FFD966"   # ô nhập liệu (yellow input)
C_GREEN  = "E2EFDA"   # ô tính tự động
C_LIGHT  = "EEF3FA"   # row alternate
C_WHITE  = "FFFFFF"
C_RED    = "C0392B"   # cảnh báo
C_TOTAL  = "D6E4F7"   # dòng tổng

# ── Styles helper ─────────────────────────────────────────
def hdr(text, bold=True, size=11, color="FFFFFF", bg=C_NAVY, wrap=False, align="center"):
    return dict(value=text, font=Font(bold=bold, size=size, color=color, name="Arial"),
                fill=PatternFill("solid", fgColor=bg),
                alignment=Alignment(horizontal=align, vertical="center", wrap_text=wrap))

def inp(text="", bold=False, bg=C_GOLD, align="left"):
    return dict(value=text, font=Font(bold=bold, size=10, name="Arial"),
                fill=PatternFill("solid", fgColor=bg),
                alignment=Alignment(horizontal=align, vertical="center"))

def calc(formula, bold=False, bg=C_GREEN, fmt=None):
    return dict(value=formula, font=Font(bold=bold, size=10, name="Arial"),
                fill=PatternFill("solid", fgColor=bg),
                alignment=Alignment(horizontal="center", vertical="center"),
                number_format=fmt or '#,##0')

def label(text, bold=False, bg=C_WHITE, align="left", size=10, color="1A1A1A"):
    return dict(value=text, font=Font(bold=bold, size=size, name="Arial", color=color),
                fill=PatternFill("solid", fgColor=bg),
                alignment=Alignment(horizontal=align, vertical="center", wrap_text=True))

def total_row(text, formula, bg=C_TOTAL):
    return (label(text, bold=True, bg=bg, align="left"),
            dict(value=formula, font=Font(bold=True, size=11, name="Arial", color=C_NAVY),
                 fill=PatternFill("solid", fgColor=bg),
                 alignment=Alignment(horizontal="center", vertical="center"),
                 number_format='#,##0'))

def apply(ws, row, col, d):
    cell = ws.cell(row=row, column=col)
    cell.value = d.get("value", "")
    if "font" in d: cell.font = d["font"]
    if "fill" in d: cell.fill = d["fill"]
    if "alignment" in d: cell.alignment = d["alignment"]
    if "number_format" in d: cell.number_format = d["number_format"]
    return cell

def thin_border(ws, r1, c1, r2, c2):
    thin = Side(style="thin", color="BBBBBB")
    for row in range(r1, r2+1):
        for col in range(c1, c2+1):
            cell = ws.cell(row=row, column=col)
            sides = {}
            if row == r1: sides['top'] = thin
            if row == r2: sides['bottom'] = thin
            if col == c1: sides['left'] = thin
            if col == c2: sides['right'] = thin
            sides['bottom'] = thin
            sides['right'] = thin
            cell.border = Border(**{k: v for k, v in sides.items()})

# ── Tạo 1 sheet PTG ───────────────────────────────────────
def make_sheet(wb, sheet_name, phuong_thuc, ck_tc_pct, ck_tts_pct, tts_label, payment_rows):
    """
    phuong_thuc : "Không Vay" | "Có Vay" | "TTS 50%" | "TTS 70%" | "TTS 95%"
    ck_tc_pct   : CK tài chính (Không vay=5%, Có vay=0%)
    ck_tts_pct  : CK TTS thêm (0 nếu không TTS)
    tts_label   : "" | "TTS 50%" | "TTS 70%" | "TTS 95%"
    payment_rows: list of (stt, label, pct_str, note) — tiến độ thanh toán
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Column widths
    col_w = [4, 42, 18, 18, 14, 12, 18]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 16

    # ── HEADER ────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    apply(ws, 1, 1, hdr("IQI VIỆT NAM  ·  SUN SYMPHONY RESIDENCE 5", size=9,
                         color="AABBCC", bg=C_NAVY, align="center"))

    ws.merge_cells("A2:G2")
    apply(ws, 2, 1, hdr(f"PHIẾU TÍNH GIÁ TẠM TÍNH — {phuong_thuc.upper()}", size=14,
                         bg=C_NAVY))

    ws.merge_cells("A3:G3")
    note_text = "⚠  Bảng tính mang tính tham khảo · Giá chính thức theo HĐMB từ Chủ Đầu Tư"
    apply(ws, 3, 1, hdr(note_text, size=8, color="FFDD99", bg=C_NAVY))

    # ── THÔNG TIN CĂN HỘ ──────────────────────────────────
    r = 5
    ws.merge_cells(f"A{r}:G{r}")
    apply(ws, r, 1, hdr("THÔNG TIN CĂN HỘ", size=10, bg=C_ORANGE))

    rows_info = [
        ("1", "Họ và tên khách hàng", "", ""),
        ("2", "Tòa", "Symphony 5", ""),
        ("3", "Tầng", "", ""),
        ("4", "Căn số", "", ""),
        ("5", "Mã Căn Hộ", "", ""),
        ("6", "Loại hình Căn Hộ", "", "Studio / 1PN+ / 2PN"),
        ("7", "Diện tích thông thủy (m²)", "", ""),
    ]

    for i, (stt, lbl, default, hint) in enumerate(rows_info):
        rr = r + 1 + i
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        ws.merge_cells(f"A{rr}:A{rr}")
        apply(ws, rr, 1, label(stt, bg=bg, align="center"))
        ws.merge_cells(f"B{rr}:D{rr}")
        apply(ws, rr, 2, label(lbl, bold=True, bg=bg))
        ws.merge_cells(f"E{rr}:G{rr}")
        cell = apply(ws, rr, 5, inp(default))
        if hint:
            ws.cell(rr, 5).comment = None  # placeholder
        ws.row_dimensions[rr].height = 20

    # ── THÔNG TIN GIÁ & CHÍNH SÁCH ────────────────────────
    r2 = r + len(rows_info) + 2
    ws.merge_cells(f"A{r2}:G{r2}")
    apply(ws, r2, 1, hdr("THÔNG TIN GIÁ & CHÍNH SÁCH", size=10, bg=C_NAVY))

    # Sub-header
    r2 += 1
    hdrs = ["STT", "Mục", "Trạng thái", "Áp dụng tại", "Thời điểm", "Mức CK (%)", "Giá trị (VND)"]
    for c, h in enumerate(hdrs, 1):
        apply(ws, r2, c, hdr(h, size=9, bg="2A4E7A"))
    ws.row_dimensions[r2].height = 20

    # Rows
    # Dòng 8: Giá niêm yết (gồm VAT & KPBT) — INPUT
    r3 = r2 + 1
    ws.row_dimensions[r3].height = 22
    apply(ws, r3, 1, label("8", align="center"))
    ws.merge_cells(f"B{r3}:D{r3}")
    apply(ws, r3, 2, label("Tổng Giá bán (đã gồm VAT & KPBT) trước CK — Niêm yết", bold=True))
    ws.merge_cells(f"E{r3}:F{r3}")
    apply(ws, r3, 5, inp("← ĐIỀN VÀO ĐÂY", align="center"))
    apply(ws, r3, 7, inp(0, align="center"))
    ws.cell(r3, 7).number_format = '#,##0'

    NIEM_YET = f"G{r3}"  # tham chiếu giá niêm yết

    # Dòng 9: Giá chưa VAT & KPBT
    r4 = r3 + 1
    ws.row_dimensions[r4].height = 20
    apply(ws, r4, 1, label("9", align="center", bg=C_GREEN))
    ws.merge_cells(f"B{r4}:D{r4}")
    apply(ws, r4, 2, label("Tổng Giá bán (chưa gồm VAT & KPBT) trước CK", bg=C_GREEN))
    ws.merge_cells(f"E{r4}:F{r4}")
    apply(ws, r4, 5, label("— tự tính", bg=C_GREEN, align="center"))
    apply(ws, r4, 7, calc(f"={NIEM_YET}/1.12"))

    CHUA_VAT = f"G{r4}"

    # Dòng 10: Early Bird — INPUT %
    r5 = r4 + 1
    ws.row_dimensions[r5].height = 20
    apply(ws, r5, 1, label("10", align="center", bg=C_LIGHT))
    ws.merge_cells(f"B{r5}:B{r5}")
    apply(ws, r5, 2, label("Chiết khấu Early Bird", bg=C_LIGHT))
    apply(ws, r5, 3, inp("Có thời hạn", align="center"))
    apply(ws, r5, 4, label("HĐTHNV", bg=C_LIGHT, align="center"))
    apply(ws, r5, 5, label("HĐTHNV", bg=C_LIGHT, align="center"))
    apply(ws, r5, 6, inp(0.05, align="center"))
    ws.cell(r5, 6).number_format = '0.0%'
    apply(ws, r5, 7, calc(f"={CHUA_VAT}*F{r5}"))

    EB_VAL = f"G{r5}"
    EB_PCT = f"F{r5}"

    # Dòng 11: CK Tài chính
    r6 = r5 + 1
    ws.row_dimensions[r6].height = 20
    apply(ws, r6, 1, label("11", align="center"))
    ws.merge_cells(f"B{r6}:B{r6}")
    apply(ws, r6, 2, label(f"Chính sách tài chính — {phuong_thuc}"))
    apply(ws, r6, 3, inp(phuong_thuc, align="center"))
    apply(ws, r6, 4, label("HĐTHNV", align="center"))
    apply(ws, r6, 5, label("HĐTHNV", align="center"))
    apply(ws, r6, 6, inp(ck_tc_pct, align="center"))
    ws.cell(r6, 6).number_format = '0.0%'
    apply(ws, r6, 7, calc(f"={CHUA_VAT}*F{r6}"))

    TC_VAL = f"G{r6}"
    TC_PCT = f"F{r6}"

    # Dòng 12: CK TTS (nếu có)
    r7 = r6 + 1
    ws.row_dimensions[r7].height = 20
    apply(ws, r7, 1, label("12", align="center", bg=C_LIGHT))
    ws.merge_cells(f"B{r7}:B{r7}")
    tts_title = f"Chiết khấu Thanh Toán Sớm — {tts_label}" if tts_label else "Chiết khấu TTS (không áp dụng)"
    apply(ws, r7, 2, label(tts_title, bg=C_LIGHT))
    apply(ws, r7, 3, inp(tts_label if tts_label else "Không áp dụng", align="center"))
    apply(ws, r7, 4, label("HĐTHNV", bg=C_LIGHT, align="center"))
    apply(ws, r7, 5, label("HĐTHNV", bg=C_LIGHT, align="center"))
    apply(ws, r7, 6, inp(ck_tts_pct, align="center"))
    ws.cell(r7, 6).number_format = '0.0%'
    apply(ws, r7, 7, calc(f"={CHUA_VAT}*F{r7}", bg=C_LIGHT if not tts_label else C_GREEN))

    TTS_VAL = f"G{r7}"
    TTS_PCT = f"F{r7}"

    # Dòng 13: CK thân thiết
    r8 = r7 + 1
    ws.row_dimensions[r8].height = 20
    apply(ws, r8, 1, label("13", align="center"))
    ws.merge_cells(f"B{r8}:B{r8}")
    apply(ws, r8, 2, label("Chiết khấu KH thân thiết (mua căn thứ 2)"))
    apply(ws, r8, 3, inp("Chưa kích hoạt", align="center"))
    apply(ws, r8, 4, label("HĐMB", align="center"))
    apply(ws, r8, 5, label("HĐMB", align="center"))
    apply(ws, r8, 6, inp(0.00, align="center"))
    ws.cell(r8, 6).number_format = '0.0%'
    apply(ws, r8, 7, calc(f"={CHUA_VAT}*F{r8}"))

    TT_VAL = f"G{r8}"

    # ── Giá sau CK (chưa VAT) ─────────────────────────────
    r9 = r8 + 1
    ws.row_dimensions[r9].height = 22
    ws.merge_cells(f"A{r9}:F{r9}")
    apply(ws, r9, 1, label("14  Tổng Giá bán (chưa VAT & KPBT) sau tất cả CK",
                            bold=True, bg=C_TOTAL, size=10))
    apply(ws, r9, 7, calc(f"={CHUA_VAT}-{EB_VAL}-{TC_VAL}-{TTS_VAL}-{TT_VAL}",
                           bold=True, bg=C_TOTAL))
    SAU_CK = f"G{r9}"

    # VAT 10%
    r10 = r9 + 1
    ws.row_dimensions[r10].height = 20
    apply(ws, r10, 1, label("15", align="center", bg=C_LIGHT))
    ws.merge_cells(f"B{r10}:E{r10}")
    apply(ws, r10, 2, label("Thuế VAT (tạm tính)", bg=C_LIGHT))
    apply(ws, r10, 6, label("10,0%", bg=C_LIGHT, align="center"))
    apply(ws, r10, 7, calc(f"={SAU_CK}*0.10", bg=C_LIGHT))
    VAT_VAL = f"G{r10}"

    # KPBT 2%
    r11 = r10 + 1
    ws.row_dimensions[r11].height = 20
    apply(ws, r11, 1, label("16", align="center"))
    ws.merge_cells(f"B{r11}:E{r11}")
    apply(ws, r11, 2, label("Kinh phí bảo trì (tạm tính)"))
    apply(ws, r11, 6, label("2,0%", align="center"))
    apply(ws, r11, 7, calc(f"={SAU_CK}*0.02"))
    KPBT_VAL = f"G{r11}"

    # TỔNG GIÁ CUỐI
    r12 = r11 + 1
    ws.merge_cells(f"A{r12}:G{r12}")
    ws.row_dimensions[r12].height = 4  # spacer

    r13 = r12 + 1
    ws.row_dimensions[r13].height = 28
    ws.merge_cells(f"A{r13}:F{r13}")
    apply(ws, r13, 1, hdr(
        "17  TỔNG GIÁ THANH TOÁN (đã gồm VAT & KPBT) — Giá trị KH phải thanh toán",
        size=11, bg=C_ORANGE))
    apply(ws, r13, 7, dict(
        value=f"={SAU_CK}+{VAT_VAL}+{KPBT_VAL}",
        font=Font(bold=True, size=13, color=C_NAVY, name="Arial"),
        fill=PatternFill("solid", fgColor="FFF0D6"),
        alignment=Alignment(horizontal="center", vertical="center"),
        number_format='#,##0'
    ))
    TONG_GIA = f"G{r13}"

    # ── TIẾN ĐỘ THANH TOÁN ────────────────────────────────
    r_td = r13 + 2
    ws.merge_cells(f"A{r_td}:G{r_td}")
    apply(ws, r_td, 1, hdr("TIẾN ĐỘ THANH TOÁN", size=10, bg=C_NAVY))

    # Date input rows
    r_td += 1
    ws.row_dimensions[r_td].height = 20
    ws.merge_cells(f"A{r_td}:B{r_td}")
    apply(ws, r_td, 1, label("Ngày ký HĐTHNV:", bold=True, bg=C_LIGHT))
    ws.merge_cells(f"C{r_td}:D{r_td}")
    apply(ws, r_td, 3, inp("", align="center"))
    apply(ws, r_td, 5, label("HĐMB dự kiến:", bold=True, bg=C_LIGHT, align="right"))
    ws.merge_cells(f"F{r_td}:G{r_td}")
    apply(ws, r_td, 6, inp("", align="center"))

    r_td += 1
    ws.row_dimensions[r_td].height = 20
    ws.merge_cells(f"A{r_td}:B{r_td}")
    apply(ws, r_td, 1, label("Bàn giao dự kiến:", bold=True, bg=C_LIGHT))
    ws.merge_cells(f"C{r_td}:D{r_td}")
    apply(ws, r_td, 3, inp("2028", align="center"))
    ws.merge_cells(f"E{r_td}:G{r_td}")
    apply(ws, r_td, 5, label("", bg=C_LIGHT))

    # Payment table header
    r_td += 1
    pay_hdrs = ["STT", "Tiến độ thanh toán", "% TT", "", "Thời gian dự kiến", "", "Số tiền"]
    for c, h in enumerate(pay_hdrs, 1):
        apply(ws, r_td, c, hdr(h, size=9, bg="2A4E7A"))
    ws.row_dimensions[r_td].height = 20

    # Payment rows
    for i, (stt, lbl_text, pct_str, note) in enumerate(payment_rows):
        rr = r_td + 1 + i
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        ws.row_dimensions[rr].height = 20
        apply(ws, rr, 1, label(str(stt) if stt else "", bg=bg, align="center"))
        ws.merge_cells(f"B{rr}:B{rr}")
        apply(ws, rr, 2, label(lbl_text, bg=bg))
        apply(ws, rr, 3, label(pct_str, bg=bg, align="center"))
        ws.merge_cells(f"D{rr}:D{rr}")
        apply(ws, rr, 4, label(note, bg=bg, size=8))
        ws.merge_cells(f"E{rr}:F{rr}")
        apply(ws, rr, 5, inp("", align="center"))  # ngày điền tay

        # Số tiền — tính theo % * tổng giá (trừ cọc)
        if pct_str and pct_str not in ("Cọc", "100% KPBT+VAT 5%", "Lãi TTS"):
            try:
                pct_val = float(pct_str.replace("%","").strip()) / 100
                apply(ws, rr, 7, calc(f"={TONG_GIA}*{pct_val}", bg=bg))
            except:
                apply(ws, rr, 7, inp("", align="center"))
        elif pct_str == "Cọc":
            apply(ws, rr, 7, inp(100000000, align="center"))
            ws.cell(rr, 7).number_format = '#,##0'
        elif pct_str == "100% KPBT+VAT 5%":
            apply(ws, rr, 7, calc(f"={KPBT_VAL}+{SAU_CK}*0.05*0.10", bg=bg))
        elif pct_str == "Lãi TTS":
            apply(ws, rr, 7, label("— xem chính sách", bg=bg, align="center", size=8))
        else:
            apply(ws, rr, 7, inp("", align="center"))

    # Dòng tổng
    r_sum = r_td + 1 + len(payment_rows)
    ws.row_dimensions[r_sum].height = 24
    ws.merge_cells(f"A{r_sum}:F{r_sum}")
    apply(ws, r_sum, 1, hdr("TỔNG GIÁ TRỊ THANH TOÁN", size=11, bg=C_NAVY))
    apply(ws, r_sum, 7, dict(
        value=f"={TONG_GIA}",
        font=Font(bold=True, size=12, color="FFFFFF", name="Arial"),
        fill=PatternFill("solid", fgColor=C_NAVY),
        alignment=Alignment(horizontal="center", vertical="center"),
        number_format='#,##0'
    ))

    # ── GHI CHÚ ───────────────────────────────────────────
    r_note = r_sum + 2
    ws.merge_cells(f"A{r_note}:G{r_note}")
    note_text = ("★  Ô màu vàng = điền thông tin  ·  Ô màu xanh = tự tính  ·  "
                 "Giá bán sau CK = Giá niêm yết ÷ 1,12 rồi trừ các chiết khấu  ·  "
                 "VAT 10% + KPBT 2% tính trên giá sau CK")
    apply(ws, r_note, 1, label(note_text, size=8, color="888888", bg=C_WHITE))

    return ws


# ── Định nghĩa tiến độ từng phương thức ──────────────────

# Tiến độ chuẩn (Có vay / Không vay) — 19 đợt ~5%/đợt
def td_tiendo(tong_dot=13):
    rows = [
        (1, "Ký HĐTHNV — Thanh toán lần 1 (Đặt cọc)", "Cọc", ""),
        (2, "Thanh toán lần 2", "15%", ""),
    ]
    dot = 3
    for i in range(tong_dot - 2):
        rows.append((dot, f"Thanh toán lần {dot}", "5%", ""))
        dot += 1
    rows.append(("", "Nhận bàn giao sử dụng (dự kiến)", "", "Sun Early Key"))
    rows.append((dot, f"Thanh toán lần {dot}", "5%", ""))
    dot += 1
    for i in range(4):
        rows.append((dot, f"Thanh toán lần {dot}", "5%", ""))
        dot += 1
    rows.append((dot, "Thanh toán lần cuối (theo TB Bàn Giao)", "100% KPBT+VAT 5%", ""))
    rows.append(("", "Lãi TTS 8%/năm (từ bàn giao → nhận GCN)", "Lãi TTS", ""))
    return rows

def td_tts95():
    return [
        (1, "Ký HĐTHNV — Thanh toán lần 1 (Đặt cọc)", "Cọc", ""),
        (2, "Thanh toán lần 2", "15%", ""),
        (3, "Thanh toán lần 3 — TTS 80% giá trị căn", "80%", "Trong vòng 10 ngày ký HĐTHNV"),
        ("", "Ký HĐMB (dự kiến)", "", ""),
        (4, "Thanh toán lần cuối (theo TB Bàn Giao)", "100% KPBT+VAT 5%", ""),
        ("", "Lãi TTS 8%/năm (từ bàn giao → nhận GCN)", "Lãi TTS", ""),
    ]

def td_tts70():
    return [
        (1, "Ký HĐTHNV — Thanh toán lần 1 (Đặt cọc)", "Cọc", ""),
        (2, "Thanh toán lần 2", "15%", ""),
        (3, "Thanh toán lần 3 — TTS 55% giá trị căn", "55%", ""),
        ("", "Ký HĐMB (dự kiến)", "", ""),
        (4, "Thanh toán lần 4", "5%", ""),
        (5, "Thanh toán lần 5", "5%", ""),
        (6, "Thanh toán lần 6", "5%", ""),
        (7, "Thanh toán lần 7", "5%", ""),
        ("", "Nhận bàn giao sử dụng", "", "Sun Early Key"),
        (8, "Thanh toán lần 8", "5%", ""),
        (9, "Thanh toán lần 9", "5%", ""),
        (10, "Thanh toán lần 10", "5%", ""),
        (11, "Thanh toán lần 11", "5%", ""),
        (12, "Thanh toán lần cuối (theo TB Bàn Giao)", "100% KPBT+VAT 5%", ""),
        ("", "Lãi TTS 8%/năm (từ bàn giao → nhận GCN)", "Lãi TTS", ""),
    ]

def td_tts50():
    return [
        (1, "Ký HĐTHNV — Thanh toán lần 1 (Đặt cọc)", "Cọc", ""),
        (2, "Thanh toán lần 2", "15%", ""),
        (3, "Thanh toán lần 3 — TTS 35% giá trị căn", "35%", ""),
        ("", "Ký HĐMB (dự kiến)", "", ""),
        (4, "Thanh toán lần 4", "5%", ""),
        (5, "Thanh toán lần 5", "5%", ""),
        (6, "Thanh toán lần 6", "5%", ""),
        (7, "Thanh toán lần 7", "5%", ""),
        (8, "Thanh toán lần 8", "5%", ""),
        (9, "Thanh toán lần 9", "5%", ""),
        (10, "Thanh toán lần 10", "5%", ""),
        (11, "Thanh toán lần 11", "5%", ""),
        ("", "Nhận bàn giao sử dụng", "", "Sun Early Key"),
        (12, "Thanh toán lần 12", "5%", ""),
        (13, "Thanh toán lần 13", "5%", ""),
        (14, "Thanh toán lần 14", "5%", ""),
        (15, "Thanh toán lần 15", "5%", ""),
        (16, "Thanh toán lần 16", "5%", ""),
        (17, "Thanh toán lần cuối (theo TB Bàn Giao)", "100% KPBT+VAT 5%", ""),
        ("", "Lãi TTS 8%/năm (từ bàn giao → nhận GCN)", "Lãi TTS", ""),
    ]

# ── Build workbook ────────────────────────────────────────
def build():
    wb = Workbook()
    wb.remove(wb.active)  # xoá sheet mặc định

    sheets = [
        ("Thanh Toán Tiến Độ", "Thanh Toán Tiến Độ (Không Vay)", 0.05, 0.00, "", td_tiendo(13)),
        ("Có Vay",     "Có Vay",          0.00, 0.00, "",       td_tiendo(8)),
        ("TTS 95%",    "TTS 95%",         0.05, 0.165, "TTS 95%", td_tts95()),
        ("TTS 70%",    "TTS 70%",         0.05, 0.09, "TTS 70%",  td_tts70()),
        ("TTS 50%",    "TTS 50%",         0.05, 0.055, "TTS 50%", td_tts50()),
    ]

    for tab_name, phuong_thuc, ck_tc, ck_tts, tts_lbl, td_rows in sheets:
        make_sheet(wb, tab_name, phuong_thuc, ck_tc, ck_tts, tts_lbl, td_rows)

    out = "/Users/macos/Desktop/content-creation/outputs/PTG-Symphony5-Template.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    return out

if __name__ == "__main__":
    build()
